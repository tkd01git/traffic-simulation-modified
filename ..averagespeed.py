import pandas as pd
import openpyxl

def column_to_letter(col):
    letter = ""
    while col >= 0:
        letter = chr(col % 26 + 65) + letter
        col = col // 26 - 1
    return letter

def process_file(file_path, target_source_file=None):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook["Sheet1"]

    # Find detector1 column
    detector_col_index = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == "detector1":
            detector_col_index = col[0].column - 1  # Convert to 0-based index
            break

    if detector_col_index is None:
        raise ValueError("detector1 not found")

    if detector_col_index >= 2:
        end_col = column_to_letter(detector_col_index - 2)
    else:
        raise ValueError("Two columns before 'detector1' do not exist.")

    df = pd.read_excel(file_path, sheet_name="Sheet1", usecols=f'A:{end_col}', header=None)
    time_data = df.iloc[0].values
    speed_data = df.iloc[1:41].values

    back_speed = speed_data[:20].mean(axis=0)
    front_speed = speed_data[20:40].mean(axis=0)
    speed_diff = front_speed - back_speed  # Calculate difference (Front - Back)

    if target_source_file:
        target_workbook = openpyxl.load_workbook(target_source_file, data_only=True)
        target_sheet = target_workbook["Sheet1"]
        target_df = pd.read_excel(target_source_file, sheet_name="Sheet1", usecols=f'A:{end_col}', header=None)
        target_avg_speed = target_df.iloc[99].values if len(target_df) > 98 else [0] * len(time_data)
    else:
        target_avg_speed = df.iloc[99].values if len(df) > 98 else [0] * len(time_data)

    return time_data, speed_diff, target_avg_speed

def find_negative_intervals(speed_diff, threshold):
    intervals = {}
    count = 0
    for i, diff in enumerate(speed_diff):
        if diff < -10:
            count += 1
        else:
            count = 0

        if count >= threshold:
            if threshold not in intervals:
                intervals[threshold] = []
            intervals[threshold].append(i - threshold + 1)  # Start index of the interval

    return intervals

def analyze_conditions(file_pairs, thresholds):
    # Initialize results
    pro_results = {t: {"t_exists_yellow_before": 0, "t_exists_no_yellow_before": 0,
                       "t_not_exists_yellow_before": 0, "t_not_exists_no_yellow_before": 0} for t in thresholds}

    repro_results = {t: {"t_exists_yellow_before": 0, "t_exists_no_yellow_before": 0,
                         "t_not_exists_yellow_before": 0, "t_not_exists_no_yellow_before": 0} for t in thresholds}

    for idx, (pro_file, repro_file) in enumerate(file_pairs):
        try:
            # Process prodata and reprodata
            pro_time, pro_speed_diff, pro_target_avg = process_file(pro_file)
            repro_time, repro_speed_diff, repro_target_avg = process_file(repro_file, target_source_file=pro_file)

            # Cache intervals for all thresholds
            pro_intervals = {t: find_negative_intervals(pro_speed_diff, t) for t in thresholds}
            repro_intervals = {t: find_negative_intervals(repro_speed_diff, t) for t in thresholds}

            for t in thresholds:
                # Analyze prodata
                pro_t_below_50 = [time for time, speed in zip(pro_time, pro_target_avg) if speed < 50]
                pro_has_t = len(pro_t_below_50) > 0
                pro_has_yellow_before_t = any(time < pro_t_below_50[0] for time in pro_intervals[t]) if pro_has_t else False

                if pro_has_t and pro_has_yellow_before_t:
                    pro_results[t]["t_exists_yellow_before"] += 1
                elif pro_has_t and not pro_has_yellow_before_t:
                    pro_results[t]["t_exists_no_yellow_before"] += 1
                elif not pro_has_t and len(pro_intervals[t]) > 0:
                    pro_results[t]["t_not_exists_yellow_before"] += 1
                else:
                    pro_results[t]["t_not_exists_no_yellow_before"] += 1

                # Analyze reprodata
                repro_t_below_50 = [time for time, speed in zip(repro_time, repro_target_avg) if speed < 50]
                repro_has_t = len(repro_t_below_50) > 0
                repro_has_yellow_before_t = any(time < repro_t_below_50[0] for time in repro_intervals[t]) if repro_has_t else False

                if repro_has_t and repro_has_yellow_before_t:
                    repro_results[t]["t_exists_yellow_before"] += 1
                elif repro_has_t and not repro_has_yellow_before_t:
                    repro_results[t]["t_exists_no_yellow_before"] += 1
                elif not repro_has_t and len(repro_intervals[t]) > 0:
                    repro_results[t]["t_not_exists_yellow_before"] += 1
                else:
                    repro_results[t]["t_not_exists_no_yellow_before"] += 1

        except Exception as e:
            print(f"Error processing file pair index {idx}: {e}")

    return pro_results, repro_results

# File pairs to process
file_pairs = [(f'prodata{i}.xlsx', f'reprodata{i}.xlsx') for i in range(1, 101)]

# Thresholds to analyze
thresholds = [3,4,5,6]

# Analyze and display results
pro_results, repro_results = analyze_conditions(file_pairs, thresholds)

print("Prodata Results:")
for t, results in pro_results.items():
    print(f"\nThreshold: {t}")
    for condition, count in results.items():
        print(f"{condition}: {count}")

print("\nReprodata Results:")
for t, results in repro_results.items():
    print(f"\nThreshold: {t}")
    for condition, count in results.items():
        print(f"{condition}: {count}")
