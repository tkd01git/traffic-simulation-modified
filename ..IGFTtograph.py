import pandas as pd
import numpy as np
from scipy.linalg import eigh

# ファイル名リストを受け取る
file_numbers =  [94]
# ここに処理するファイルの番号をリストで指定
userank = 2

# Laplacian行列の作成
def create_laplacian(n):
    A = np.zeros((n, n))
    for i in range(n - 1):
        A[i, i + 1] = 1
        A[i + 1, i] = 1
    L = np.diag(np.sum(A, axis=1)) - A
    return np.dot(
        np.linalg.inv(np.sqrt(np.diag(np.sum(A, axis=1)))),
        np.dot(L, np.linalg.inv(np.sqrt(np.diag(np.sum(A, axis=1)))))
    )

# メイン処理
def process_file(input_file, output_file):
    # データ読み込み
    data = pd.read_excel(input_file, sheet_name="Sheet1", header=None)
    time_data = data.iloc[0, 0:].values  # 時刻データ
    speed_data = data.iloc[1:41, 0:].values  # 速度データ
    row_100_data = data.iloc[99].values  # 100行目のデータ

    # グラフのノード数
    n = 40

    # Laplacian行列を準備
    L_normalized = create_laplacian(n)
    eigenvalues, eigenvectors = eigh(L_normalized)

    # 固有値と固有ベクトルをソート
    sorted_indices = np.argsort(eigenvalues)
    sorted_eigenvalues = eigenvalues[sorted_indices]
    sorted_eigenvectors = eigenvectors[:, sorted_indices]

    # 再構成データを格納するリスト
    reconstructed_data = []

    for t in range(speed_data.shape[1]):
        # 各時刻の速度データ
        speed_at_t = speed_data[:, t]

        # フーリエ係数を計算
        F_lambda = np.array([np.dot(speed_at_t, sorted_eigenvectors[:, i]) for i in range(n)])

        # 支配的な3モードを抽出
        dominant_indices = np.argsort(np.abs(F_lambda))[-userank:][::-1]  # 絶対値が大きい上位5つ

        # 再構成
        reconstructed_signal = sum(F_lambda[i] * sorted_eigenvectors[:, i] for i in dominant_indices)
        reconstructed_data.append(reconstructed_signal)

    # 再構成データをデータフレームに変換
    reconstructed_data = np.array(reconstructed_data).T
    output_df = pd.DataFrame(reconstructed_data, columns=time_data)

    # 100行目のデータをデータフレームに追加
    output_df.loc[99] = row_100_data  # データフレームの100行目にデータを追加

    # Excelファイルに書き込む
    output_df.to_excel(output_file, index=False, header=True)
    print(f"再構成されたデータが {output_file} に保存されました。")

# 指定されたファイルを順次処理
for num in file_numbers:
    input_file = f'prodata{num}.xlsx'
    output_file = f'reprodata{num}.xlsx'
    process_file(input_file, output_file)

import pandas as pd
import numpy as np
from scipy.linalg import eigh
import openpyxl


# ファイルを順に処理
file_indices = file_numbers
input_files = [f"reprodata{x}.xlsx" for x in file_indices]


# Excel列番号をアルファベットに変換する関数
def column_to_letter(column_index):
    letter = ""
    while column_index >= 0:
        letter = chr(column_index % 26 + ord('A')) + letter
        column_index = column_index // 26 - 1
    return letter

# Laplacian行列の作成
def create_laplacian(n):
    A = np.zeros((n, n))
    for i in range(n - 1):
        A[i, i + 1] = 1
        A[i + 1, i] = 1
    L = np.diag(np.sum(A, axis=1)) - A
    return np.dot(
        np.linalg.inv(np.sqrt(np.diag(np.sum(A, axis=1)))),
        np.dot(L, np.linalg.inv(np.sqrt(np.diag(np.sum(A, axis=1)))))
    )

# Fλの実部を計算する関数
def calculate_F_lambda_real_parts(df, L_normalized):
    real_parts_list = []
    for column in df.columns:
        speed_data = df[column].values.flatten()
        eigenvalues, eigenvectors = eigh(L_normalized, eigvals_only=False)

        sorted_indices = np.argsort(eigenvalues)
        sorted_eigenvectors = eigenvectors[:, sorted_indices]

        F_lambda = np.zeros(sorted_eigenvectors.shape[1], dtype=complex)
        for i in range(sorted_eigenvectors.shape[1]):
            u_lambda_i = sorted_eigenvectors[:, i]
            F_lambda[i] = sum(speed_data * np.conj(u_lambda_i))

        real_parts = F_lambda.real
        real_parts_list.append(real_parts)
    return real_parts_list

# Fλ のランキングを計算する関数
def calculate_F_lambda_global_rank(df, L_normalized):
    global_ranks = []
    for column in df.columns:
        speed_data = df[column].values.flatten()
        eigenvalues, eigenvectors = eigh(L_normalized, eigvals_only=False)

        sorted_indices = np.argsort(eigenvalues)
        sorted_eigenvectors = eigenvectors[:, sorted_indices]

        # 各固有値に対応するフーリエ係数 Fλ を計算
        F_lambda = np.zeros(sorted_eigenvectors.shape[1], dtype=complex)
        for i in range(sorted_eigenvectors.shape[1]):
            u_lambda_i = sorted_eigenvectors[:, i]
            F_lambda[i] = sum(speed_data * np.conj(u_lambda_i))

        # Fλ の絶対値を計算し、ランキングを決定
        abs_F_lambda = np.abs(F_lambda)
        rank = abs_F_lambda.argsort().argsort()  # 小さい順のランキングを取得
        global_ranks.append(rank + 1)  # 1位が最大値になるよう反転（1位が最小値にならないように）

    return global_ranks

# シートにデータを書き込む関数
def write_sheet_with_ranks(writer, sheet_name, time_data, speed_data, real_parts_list):
    output_data = {
        'Time': list(time_data),
        'speed': list(speed_data)}
    for lambda_index in range(40):
        output_data[f'F_lambda{lambda_index+1}'] = [
            real_parts[lambda_index] if len(real_parts) > lambda_index else '' for real_parts in real_parts_list
        ]

    results_df = pd.DataFrame(output_data)
    results_df.to_excel(writer, index=False, sheet_name=sheet_name)


# Laplacian行列を準備
n = 40
L_normalized = create_laplacian(n)


# ファイルを順に処理
for file_index, input_file in zip(file_indices, input_files):
    result_file1 = f"reresult{file_index}.xlsx"
    result_file2 = f"reλrankresult{file_index}.xlsx"

    # ファイル読み込み
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook["Sheet1"]

    detector_col_index = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == "detector1":
            detector_col_index = col[0].column - 1  # 0ベースに変換
            break

    if detector_col_index is not None:
        detector_col_letter = column_to_letter(detector_col_index)
        if detector_col_index >= 2:
            two_columns_before_letter = column_to_letter(detector_col_index - 2)
        else:
            raise ValueError("2つ前の列は存在しません。")
    else:
        raise ValueError("Detector1が見つかりません。")

    start_col = 'A'
    end_col = two_columns_before_letter
    sheet_name1 = 'Sheet1'
    sheet_name2 = 'Sheet2'

    df1 = pd.read_excel(input_file, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None)
    data_frames = [
        pd.read_excel(input_file, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=i, nrows=40)
        for i in [1]
    ]

    time_data = df1.iloc[0].values.flatten()
    speed_data = df1.iloc[41].values.flatten()

    # 1つ目の結果ファイルを作成
    with pd.ExcelWriter(result_file1, mode='w') as writer:
        for i, (df, sheet_name) in enumerate(zip(
            data_frames, 
            [f'Sheet{i+1}' for i in range(len(data_frames))]
        )):
            real_parts_list = calculate_F_lambda_real_parts(df, L_normalized)
            write_sheet_with_ranks(writer, sheet_name, time_data, speed_data, real_parts_list)

    print(f"{result_file1} が保存されました。")
    
import pandas as pd
import matplotlib.pyplot as plt
import os

# スピードの閾値
speed_threshold = 50

# 処理するファイル名のリスト
file_numbers =  [81]
base_path = "C://Users//Ytakada//Downloads//1220traffic-simulation-de"
file_names = [f"reresult{num}.xlsx" for num in file_numbers]

# 赤線を引く時刻を決定する関数
def identify_red_lines(df, speed_threshold):
    return df.loc[df['Speed'] < speed_threshold, 'Time'].tolist()

# グラフの描画準備
num_files = len(file_numbers)
fig, axes = plt.subplots(nrows=num_files, ncols=2, figsize=(10, 5 * num_files), sharex=False, sharey=False)

# 1つのファイルしかない場合、axes をリストとして扱う
if num_files == 1:
    axes = [axes]  # サブプロットをリスト化

# F_lambda2 が連続して負の値を取る時刻を取得する関数
def identify_yellow_lines(df, consecutive_threshold=5):
    """連続して負の値を取る時刻を特定"""
    negative_indices = df['F_lambda2'] < 0
    groups = negative_indices.ne(negative_indices.shift()).cumsum()
    negative_streaks = df.groupby(groups).filter(lambda g: len(g) >= consecutive_threshold and g['F_lambda2'].iloc[0] < 0)
    return negative_streaks['Time'].tolist()

# 各ファイルを処理
for row_idx, file_name in enumerate(file_names):
    file_path = os.path.join(base_path, file_name)
    file_number = int(file_name.replace('reresult', '').replace('.xlsx', ''))

    try:
        # ファイルとシート1を読み込み
        df_sheet1 = pd.read_excel(file_path, sheet_name='Sheet1')
        df_sheet1 = df_sheet1[['Time', 'speed', 'F_lambda2']].dropna()
        df_sheet1.columns = ['Time', 'Speed', 'F_lambda2']  # 列名を統一

        # 赤線を引く時刻を決定
        thin_red_lines = identify_red_lines(df_sheet1, speed_threshold)

        # 黄色線を引く時刻を決定（連続する負の値を考慮）
        yellow_lines = identify_yellow_lines(df_sheet1, consecutive_threshold=5)

        # 速度グラフのプロット
        ax_speed = axes[row_idx][0]  # 左側の列
        ax_speed.plot(df_sheet1['Time'], df_sheet1['Speed'], label='Speed', color='purple', linewidth=1.2)

        for thin_time in thin_red_lines:  # 薄い赤線を描画
            ax_speed.axvline(x=thin_time, color='red', linestyle='-', alpha=0.5, linewidth=1.2)

        ax_speed.set_title(f'File {file_number} - Speed', fontsize=9)
        ax_speed.set_ylabel("Speed (km/h)", fontsize=8)
        ax_speed.set_xlabel("Time (s)", fontsize=8)
        ax_speed.legend(fontsize=8)
        ax_speed.grid()
        ax_speed.set_ylim(0, 100)

        # F_lambda2 グラフのプロット
        ax_frambda = axes[row_idx][1]  # 右側の列
        ax_frambda.plot(df_sheet1['Time'], df_sheet1['F_lambda2'], color='blue', linewidth=1)

        for thin_time in thin_red_lines:
            ax_frambda.axvline(x=thin_time, color='red', linestyle='-', alpha=0.5, linewidth=1.2)

        for yellow_time in yellow_lines:  # 負の値に対応する連続時刻で黄色線を描画
            ax_frambda.axvline(x=yellow_time, color='yellow', linestyle='--', alpha=0.8, linewidth=1)

        ax_frambda.set_ylim(-120, 120)
        ax_frambda.set_title(f'File {file_number} - F_lambda2', fontsize=9)
        ax_frambda.set_ylabel("F_lambda2", fontsize=8)
        ax_frambda.set_xlabel("Time (s)", fontsize=8)
        ax_frambda.grid()

    except Exception as e:
        print(f"Error processing file {file_name}: {e}")
        continue

# 全体のタイトルを追加
fig.suptitle("Speed and F_lambda2", fontsize=14)

# 注意書きを追加
plt.figtext(0.5, 0.01, "Red line = speed below 50 km/h. Yellow dashed line = F_lambda2 is negative for 5+ consecutive times.", ha="center", fontsize=10, color="gray")

plt.show()
