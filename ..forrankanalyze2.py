import pandas as pd
import numpy as np
from scipy.linalg import eigh
import openpyxl

def column_to_letter(column_index):
    """Excelの列番号をアルファベットに変換"""
    letter = ""
    while column_index >= 0:
        letter = chr(column_index % 26 + ord('A')) + letter
        column_index = column_index // 26 - 1
    return letter

# ファイルを順に処理
file_numbers = range(1,101)
file_indices = file_numbers
input_files = [f"reprodata{x}.xlsx" for x in file_indices]

# Laplacian行列の作成
def create_laplacian(n):
    A = np.zeros((n, n))
    for i in range(n - 1):
        A[i, i + 1] = 1
        A[i + 1, i] = 1
    L = np.diag(np.sum(A, axis=1)) - A
    return np.dot(
        np.linalg.inv(np.sqrt(np.diag(np.sum(A, axis=1)))),
        np.dot(L, np.linalg.inv(np.sqrt(np.diag(np.sum(A, axis=1)))))
    )

# Fλの実部を計算する関数
def calculate_F_lambda_real_parts(df, L_normalized):
    real_parts_list = []
    for column in df.columns:
        speed_data = df[column].values.flatten()
        eigenvalues, eigenvectors = eigh(L_normalized, eigvals_only=False)

        sorted_indices = np.argsort(eigenvalues)
        sorted_eigenvectors = eigenvectors[:, sorted_indices]

        F_lambda = np.zeros(sorted_eigenvectors.shape[1], dtype=complex)
        for i in range(sorted_eigenvectors.shape[1]):
            u_lambda_i = sorted_eigenvectors[:, i]
            F_lambda[i] = sum(speed_data * np.conj(u_lambda_i))

        real_parts = F_lambda.real
        real_parts_list.append(real_parts)
    return real_parts_list

# Frambda2 の順位を計算して結果に追加する関数
def add_F_lambda2_rank(writer, sheet_name, time_data, speed_data, real_parts_list):
    output_data = {
        'Time': list(time_data),
        'speed': list(speed_data)
    }

    # Frambda の列を追加
    for lambda_index in range(40):
        output_data[f'F_lambda{lambda_index+1}'] = [
            real_parts[lambda_index] if len(real_parts) > lambda_index else '' for real_parts in real_parts_list
        ]

    # Frambda2 の順位を計算
    F_lambda2_ranks = []
    for real_parts in real_parts_list:
        abs_values = np.abs(real_parts)
        if len(abs_values) >= 40:
            rank = abs_values.argsort()[::-1].argsort() + 1  # 降順にソートし順位付け
            F_lambda2_ranks.append(rank[1])  # F_lambda2 の順位を取得
        else:
            F_lambda2_ranks.append('')

    output_data['Rank of F_lambda2'] = F_lambda2_ranks

    # 結果をデータフレームに変換して書き出し
    results_df = pd.DataFrame(output_data)
    results_df.to_excel(writer, index=False, sheet_name=sheet_name)

# Laplacian行列を準備
n = 40
L_normalized = create_laplacian(n)

# ファイルを順に処理
for file_index, input_file in zip(file_indices, input_files):
    result_file1 = f"result{file_index}.xlsx"

    # ファイル読み込み
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook["Sheet1"]

    detector_col_index = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == "detector1":
            detector_col_index = col[0].column - 1  # 0ベースに変換
            break

    if detector_col_index is not None:
        detector_col_letter = column_to_letter(detector_col_index)
        if detector_col_index >= 2:
            two_columns_before_letter = column_to_letter(detector_col_index - 2)
        else:
            raise ValueError("2つ前の列は存在しません。")
    else:
        raise ValueError("Detector1が見つかりません。")

    start_col = 'A'
    end_col = two_columns_before_letter
    sheet_name1 = 'Sheet1'
    
    df1 = pd.read_excel(input_file, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None)
    data_frames = [
        pd.read_excel(input_file, sheet_name=sheet_name1, usecols=f'{start_col}:{end_col}', header=None, skiprows=i, nrows=40)
        for i in [1]
    ]

    time_data = df1.iloc[0].values.flatten()
    speed_data = df1.iloc[99].values.flatten()

    # 1つ目の結果ファイルを作成
    with pd.ExcelWriter(result_file1, mode='w') as writer:
        for i, (df, sheet_name) in enumerate(zip(
            data_frames, 
            [f'Sheet{i+1}' for i in range(len(data_frames))]
        )):
            real_parts_list = calculate_F_lambda_real_parts(df, L_normalized)
            add_F_lambda2_rank(writer, sheet_name, time_data, speed_data, real_parts_list)

    print(f"{result_file1} が保存されました。")
